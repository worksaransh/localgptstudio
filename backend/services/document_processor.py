import os
import PyPDF2
from docx import Document as DocxDocument
import csv
import io
from typing import Optional
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook


class DocumentProcessor:
    def __init__(self):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=512,
            chunk_overlap=64,
            separators=["\n\n", "\n", ".", " ", ""],
        )

    def extract_text(self, filepath: str) -> str:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".pdf":
            return self._extract_pdf(filepath)
        elif ext == ".docx":
            return self._extract_docx(filepath)
        elif ext == ".xlsx":
            return self._extract_xlsx(filepath)
        elif ext == ".csv":
            return self._extract_csv(filepath)
        elif ext == ".txt":
            return self._extract_txt(filepath)
        return ""

    def _extract_pdf(self, filepath: str) -> str:
        text = []
        with open(filepath, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text.append(page.extract_text())
        return "\n".join(text)

    def _extract_docx(self, filepath: str) -> str:
        doc = DocxDocument(filepath)
        return "\n".join([p.text for p in doc.paragraphs])

    def _extract_xlsx(self, filepath: str) -> str:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                text.append(", ".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(text)

    def _extract_csv(self, filepath: str) -> str:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            return "\n".join([",".join(row) for row in reader])

    def _extract_txt(self, filepath: str) -> str:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    def chunk_text(self, text: str) -> list[str]:
        return self.text_splitter.split_text(text)

    def process(self, filepath: str) -> tuple[str, list[str]]:
        text = self.extract_text(filepath)
        chunks = self.chunk_text(text)
        return text, chunks
